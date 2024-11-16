##############################################################################
#
# Simple Python program to benchmark several Python Excel writing modules.
#
# python bench_excel_writers.py [num_rows] [num_cols]
#
# Copyright 2013-2024, John McNamara, jmcnamara@cpan.org
#
import os
import sys
try:
    from time import process_time
except ImportError:
    from time import clock as process_time

import openpyxl
import xlsxwriter

# Default to 1 sheet with 1000 rows x 50 cols
row_max = 1000
col_max = 50
sheets = 1

if len(sys.argv) > 1:
    row_max = int(sys.argv[1])
    sheets = int(sys.argv[3])

elif len(sys.argv) > 2:
    col_max = int(sys.argv[2])

elif len(sys.argv) > 3:
    sheets = int(sys.argv[3])


def print_elapsed_time(module_name, elapsed, optimised=False):
    """ Print module run times in a consistent format. """
    if optimised:
        module_name += " (optimised)"
    print(f"    {module_name:22s}: {elapsed:6.2f}")


def time_xlsxwriter(optimised=False):
    """ Run XlsxWriter in optimised/constant memory mode. """
    options = {}
    module_name = "xlsxwriter"
    if optimised:
        options['constant_memory'] = True

    start_time = process_time()
    filename = 'xlsxwriter_opt.xlsx'

    workbook = xlsxwriter.Workbook(filename,
                                   options=options)
    for r in range(sheets):
        worksheet = workbook.add_worksheet()

        for row in range(0, row_max, 2):
            string_data = [f"Row: {row} Col: {col}" for col in range(col_max)]
            worksheet.write_row(row, 0, string_data)

            num_data = [row + col for col in range(col_max)]
            worksheet.write_row(row + 1, 0, num_data)

    workbook.close()

    elapsed = process_time() - start_time
    print_elapsed_time(module_name, elapsed, optimised)
    os.remove(filename)


def time_openpyxl(optimised=False):
    """ Run OpenPyXL in default mode. """
    module_name = "openpyxl"

    start_time = process_time()
    filename = 'openpyxl.xlsx'

    workbook = openpyxl.Workbook(write_only=optimised)
    for r in range(sheets):
        worksheet = workbook.create_sheet()

        for row in range(row_max // 2):

            string_data = (f"Row: {row} Col: {col}" for col in range(col_max))
            worksheet.append(string_data)

            num_data = (row + col for col in range(col_max))
            worksheet.append(num_data)

    workbook.save(filename)

    elapsed = process_time() - start_time
    print_elapsed_time(module_name, elapsed, optimised)
    os.remove(filename)


print("")
print("Versions:")
print(f"python: {str(sys.version).split()[0]}")
print(f"openpyxl: {openpyxl.__version__}")
print(f"xlsxwriter: {xlsxwriter.__version__}")
print("")

print("Dimensions:")
print(f"    Rows   = {row_max}")
print(f"    Cols   = {col_max}")
print(f"    Sheets = {sheets}")
print("")

print("Times:")
time_xlsxwriter()
time_xlsxwriter(optimised=True)
time_openpyxl()
time_openpyxl(optimised=True)
print("")

