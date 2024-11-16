##############################################################################
#
# Simple Python program to benchmark several Python Excel writing modules.
#
# python bench_excel_writers.py [num_rows] [num_cols]
#
# Copyright 2013-2024, John McNamara, jmcnamara@cpan.org
#

import sys
from time import perf_counter

import openpyxl
import pyexcelerate
import xlsxwriter
import xlwt


# Default to 1000 rows x 50 cols.
if len(sys.argv) > 1:
    row_max = int(sys.argv[1])
    col_max = 50
else:
    row_max = 1000
    col_max = 50

if len(sys.argv) > 2:
    col_max = int(sys.argv[2])


def print_elapsed_time(module_name, elapsed):
    """ Print module run times in a consistent format. """
    print(f"    {module_name:28s}: {elapsed:6.2f}")


def time_xlsxwriter():
    """ Run XlsxWriter in default mode. """
    start_time = perf_counter()

    workbook = xlsxwriter.Workbook('xlsxwriter.xlsx')
    worksheet = workbook.add_worksheet()

    for row in range(row_max // 2):
        for col in range(col_max):
            worksheet.write_string(row * 2, col, f"Row: {row} Col: {col}")
        for col in range(col_max):
            worksheet.write_number(row * 2 + 1, col, row + col)

    workbook.close()

    elapsed = perf_counter() - start_time
    print_elapsed_time('xlsxwriter', elapsed)


def time_xlsxwriter_optimised():
    """ Run XlsxWriter in optimised/constant memory mode. """
    start_time = perf_counter()

    workbook = xlsxwriter.Workbook('xlsxwriter_opt.xlsx',
                                   {'constant_memory': True})
    worksheet = workbook.add_worksheet()

    for row in range(row_max // 2):
        for col in range(col_max):
            worksheet.write_string(row * 2, col, f"Row: {row} Col: {col}")
        for col in range(col_max):
            worksheet.write_number(row * 2 + 1, col, row + col)

    workbook.close()

    elapsed = perf_counter() - start_time
    print_elapsed_time('xlsxwriter (constant_memory)', elapsed)


def time_openpyxl():
    """ Run OpenPyXL in default mode. """
    start_time = perf_counter()

    workbook = openpyxl.workbook.Workbook()
    worksheet = workbook.active

    for row in range(row_max // 2):
        for col in range(col_max):
            worksheet.cell(row * 2 + 1, col + 1, f"Row: {row} Col: {col}")
        for col in range(col_max):
            worksheet.cell(row * 2 + 2, col + 1, row + col)

    workbook.save('openpyxl.xlsx')

    elapsed = perf_counter() - start_time
    print_elapsed_time('openpyxl', elapsed)


def time_openpyxl_optimised():
    """ Run OpenPyXL in optimised mode. """
    start_time = perf_counter()

    workbook = openpyxl.workbook.Workbook(write_only=True)
    worksheet = workbook.create_sheet()

    for row in range(row_max // 2):
        string_data = [f"Row: {row} Col: {col}" for col in range(col_max)]
        worksheet.append(string_data)

        num_data = [row + col for col in range(col_max)]
        worksheet.append(num_data)

    workbook.save('openpyxl_opt.xlsx')

    elapsed = perf_counter() - start_time
    print_elapsed_time('openpyxl   (optimised)', elapsed)


def time_pyexcelerate():
    """ Run pyexcelerate in "faster" mode. """
    start_time = perf_counter()

    workbook = pyexcelerate.Workbook()
    worksheet = workbook.new_sheet('Sheet1')

    for row in range(row_max // 2):
        for col in range(col_max):
            worksheet.set_cell_value(row * 2 + 1, col + 1, f"Row: {row} Col: {col}")
        for col in range(col_max):
            worksheet.set_cell_value(row * 2 + 2, col + 1, row + col)

    workbook.save('pyexcelerate.xlsx')
    elapsed = perf_counter() - start_time

    print_elapsed_time('pyexcelerate', elapsed)


def time_xlwt():
    """ Run xlwt in default mode. """
    start_time = perf_counter()

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet('Sheet1')

    for row in range(row_max // 2):
        for col in range(col_max):
            worksheet.write(row * 2, col, f"Row: {row} Col: {col}")
        for col in range(col_max):
            worksheet.write(row * 2 + 1, col, row + col)

    workbook.save('xlwt.xls')

    elapsed = perf_counter() - start_time
    print_elapsed_time('xlwt', elapsed)


print("")
print("Versions:")
print(f"    python:       {str(sys.version).split()[0]}")
print(f"    openpyxl:     {openpyxl.__version__}")
print(f"    pyexcelerate: {pyexcelerate.__version__}")
print(f"    xlsxwriter:   {xlsxwriter.__version__}")
print(f"    xlwt:         {xlwt.__VERSION__}")
print("")

print("Dimensions:")
print(f"    Rows   = {row_max}")
print(f"    Cols   = {col_max}")
print("")

print("Times:")
time_pyexcelerate()
time_xlwt()
time_xlsxwriter_optimised()
time_xlsxwriter()
time_openpyxl_optimised()
time_openpyxl()
print("")
